# Convert Pandas df to xlsx
# Eg implementation:
# df_to_xlsx( {'Excel_file':df}, filename='path/to/file.xlsx'

import re
import logging
from pandas import DataFrame
from pandas import isnull
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

logging.getLogger().setLevel(logging.INFO)

def num_rows(df):
    """Returns numbers of rows for a DataFrame."""
    if df is None or df.empty:
        return 0
    return df.shape[0]

def is_val_empty(value):
    """Helper function for checking empty values"""
    if value is None:
        return True
    if isinstance(value, str):
        if value.strip() == '':
            return True
        elif value == 'None':
            return True
        elif value == 'NaT':
            return True
        return False
    if isnull(value):
        return True
    return False

def df_to_xlsx(
    df_map, filename=None, columns_map=None, has_headers=True,
    return_workbook=False
):
    """Write a dataframe to an Excel xlsx file."""
    if filename:
        assert filename.endswith("xlsx")
        logging.info("Writing dataframe to xlsx file [{}]".format(filename))
    else:
        logging.info("Writing dataframe to in-memory workbook")

    if not df_map:
        logging.warn("[df_map] is empty, no xlsx created")
        return False

    for sheet_name, _df in df_map.items():
        logging.info('    Sheet [{}]: {} rows'.format(sheet_name, num_rows(_df)))

    if not isinstance(df_map, dict):
        raise Exception('[df_map] must be a dict {sheetname: df}')

    wb = Workbook()

    # Remove default sheet
    sheet_names = wb.get_sheet_names()
    if len(sheet_names) == 1:
        sheet1 = wb.get_sheet_by_name(sheet_names[0])
        wb.remove_sheet(sheet1)

    def _remove_bad_xlsx_vals(val):
        if is_val_empty(val):
            return None
        if not isinstance(val, str):
            return val
        val = re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', val)
        return val

    for sheetname, df in sorted(df_map.items()):
        df = df.applymap(_remove_bad_xlsx_vals)

        ws = wb.create_sheet(sheetname)
        index_name = df.index.name
        has_index = True if index_name else False

        if columns_map and sheetname in columns_map:
            df = df[columns_map[sheetname]]

        for r in dataframe_to_rows(df, index=has_index, header=has_headers):
            ws.append(r)

        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

        freeze_cell = None
        if has_index:
            ws['A1'] = index_name
            freeze_cell = 'B2' if has_headers else 'B1'
        elif has_headers:
            freeze_cell = 'A2'

        if freeze_cell:
            logging.info("Freezing cell {}".format(freeze_cell))
            ws.freeze_panes = freeze_cell
    if return_workbook:
        return wb
    wb.save(filename)
    return {sheet_name:num_rows(_df) for sheet_name, _df in
                  df_map.items()}
